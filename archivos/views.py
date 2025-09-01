from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import TemplateView, ListView, CreateView
from django.views import View
from django.http import HttpResponse, Http404, FileResponse
from django.urls import reverse_lazy
from django.db.models import Count, Q, Max
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
import mimetypes
import os
from django.core.files.base import ContentFile
from django.utils._os import safe_join
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from .models import Archivo, Fraccion, PerfilUsuario, HistorialAcceso
from .forms import ArchivoForm


class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista principal del dashboard"""
    template_name = 'archivos/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener perfil del usuario
        try:
            perfil = self.request.user.perfilusuario
            tipo_usuario = perfil.tipo_usuario
        except PerfilUsuario.DoesNotExist:
            tipo_usuario = None
            messages.warning(self.request, 'Tu usuario no tiene un perfil asignado. Contacta al administrador.')
        
        # Filtrar fracciones según el tipo de usuario
        if tipo_usuario:
            fracciones = Fraccion.objects.filter(
                tipo_usuario_asignado=tipo_usuario,
                activa=True
            )
        else:
            fracciones = Fraccion.objects.filter(activa=True)
        
        # Estadísticas básicas
        total_archivos = Archivo.objects.filter(
            fraccion__in=fracciones,
            vigente=True
        ).count()
        
        archivos_recientes = Archivo.objects.filter(
            fraccion__in=fracciones
        ).order_by('-created_at')[:5]
        
        context.update({
            'fracciones': fracciones,
            'total_archivos': total_archivos,
            'archivos_recientes': archivos_recientes,
            'tipo_usuario': tipo_usuario,
        })
        
        return context


class CargarArchivoView(LoginRequiredMixin, CreateView):
    """Vista para cargar archivos"""
    model = Archivo
    form_class = ArchivoForm
    template_name = 'archivos/cargar_archivo.html'
    success_url = reverse_lazy('archivos:dashboard')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # DEBUG: Imprimir datos recibidos
        print("=== DEBUG FORMULARIO MÚLTIPLES ARCHIVOS ===")
        print(f"POST data: {self.request.POST}")
        print(f"FILES data: {self.request.FILES}")
        print(f"Form cleaned data: {form.cleaned_data}")
        print(f"Usuario: {self.request.user}")
    
        # Verificar perfil de usuario
        try:
            perfil = self.request.user.perfilusuario
            print(f"Perfil usuario: {perfil.tipo_usuario}")
        except PerfilUsuario.DoesNotExist:
            messages.error(self.request, 'Tu usuario no tiene un perfil asignado. Contacta al administrador.')
            return self.form_invalid(form)
    
        # Obtener múltiples archivos
        archivos_subidos = self.request.FILES.getlist('archivo')
        print(f"📁 Archivos recibidos: {len(archivos_subidos)}")
    
        if not archivos_subidos:
            print("❌ No se recibieron archivos")
            messages.error(self.request, 'No se recibió ningún archivo.')
            return self.form_invalid(form)
    
        # Verificar que la fracción corresponde al usuario
        fraccion = form.cleaned_data['fraccion']
        if fraccion.tipo_usuario_asignado != perfil.tipo_usuario:
            print(f"❌ Fracción no permitida para usuario: {fraccion.tipo_usuario_asignado} != {perfil.tipo_usuario}")
            messages.error(self.request, 'No tienes permisos para cargar archivos en esta fracción.')
            return self.form_invalid(form)
    
        # PROCESAR CADA ARCHIVO
        archivos_creados = []
        archivos_con_error = []
    
           # Usar transacción atómica para todos los archivos
        try:
            with transaction.atomic():
                # 🔥 PASO 1: Marcar archivos anteriores como no vigentes ANTES de crear nuevos
                archivos_existentes = Archivo.objects.filter(
                    fraccion=form.cleaned_data['fraccion'],
                    año=form.cleaned_data['año'],
                    periodo_especifico=form.cleaned_data['periodo_especifico'],
                    vigente=True  # ← SOLO los que están vigentes
                )

                nueva_version = 1
                if archivos_existentes.exists():
                    ultima_version = archivos_existentes.aggregate(
                        max_version=Max('version')
                    )['max_version'] or 0
                    nueva_version = ultima_version + 1
                    print(f"📝 Nueva versión calculada: {nueva_version}")

                    # ⚡ MARCAR COMO NO VIGENTES ANTES DE CREAR NUEVOS
                    cantidad_marcados = archivos_existentes.update(vigente=False)
                    print(f"🔄 {cantidad_marcados} archivos marcados como no vigentes")
                else:
                    print("📝 Primera versión para esta combinación")

                # 🔥 PASO 2: CREAR TODOS LOS ARCHIVOS NUEVOS COMO VIGENTES
                archivos_creados = []
                archivos_con_error = []

                for i, archivo_file in enumerate(archivos_subidos):
                    print(f"\n--- Procesando archivo {i+1}/{len(archivos_subidos)}: {archivo_file.name} ---")

                    try:
                        # Validar cada archivo individualmente
                        if archivo_file.size > 104857600:
                            raise ValidationError(f'Archivo "{archivo_file.name}" muy grande: {archivo_file.size/1024/1024:.1f} MB')

                        if archivo_file.size == 0:
                            raise ValidationError(f'Archivo "{archivo_file.name}" está vacío')

                        # Validar extensión
                        extensiones_permitidas = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
                        nombre_archivo = archivo_file.name.lower()
                        if not any(nombre_archivo.endswith(ext) for ext in extensiones_permitidas):
                            raise ValidationError(f'Archivo "{archivo_file.name}" tiene formato no permitido')

                        # 🎯 CREAR ARCHIVO - TODOS VIGENTES
                        archivo_instance = Archivo(
                            fraccion=form.cleaned_data['fraccion'],
                            usuario=self.request.user,
                            tipo_periodo=form.cleaned_data['tipo_periodo'],
                            año=form.cleaned_data['año'],
                            periodo_especifico=form.cleaned_data['periodo_especifico'],
                            archivo=archivo_file,
                            nombre_original=archivo_file.name,
                            tamaño=archivo_file.size,
                            vigente=True,  # ✅ TODOS VIGENTES
                            version=nueva_version  # ✅ MISMA VERSIÓN PARA TODOS
                        )

                        # Guardar archivo
                        archivo_instance.save()
                        archivos_creados.append(archivo_instance)

                        print(f"✅ Archivo guardado como VIGENTE: {archivo_instance.nombre_original} (v{nueva_version})")

                    except ValidationError as e:
                        print(f"❌ Error validando {archivo_file.name}: {e}")
                        archivos_con_error.append(f"{archivo_file.name}: {e}")
                    except Exception as e:
                        print(f"❌ Error inesperado con {archivo_file.name}: {e}")
                        archivos_con_error.append(f"{archivo_file.name}: Error inesperado")

                # Si hay errores, lanzar excepción para hacer rollback
                if archivos_con_error:
                    raise ValidationError("Errores en algunos archivos")

                print(f"🎉 RESUMEN: {len(archivos_creados)} archivos creados como VIGENTES (versión {nueva_version})")

        except Exception as e:
            print(f"❌ Error en transacción: {e}")
        
            # Mostrar errores específicos
            if archivos_con_error:
                for error in archivos_con_error:
                    messages.error(self.request, f"❌ {error}")
            else:
                messages.error(self.request, f'Error al guardar archivos: {e}')
        
            return self.form_invalid(form)
    
        # MOSTRAR RESULTADO EXITOSO
        if archivos_creados:
            if len(archivos_creados) == 1:
                messages.success(
                    self.request, 
                    f'✅ Archivo "{archivos_creados[0].nombre_original}" cargado exitosamente como versión {nueva_version}.'
                )
            else:
                messages.success(
                    self.request, 
                    f'✅ {len(archivos_creados)} archivos cargados exitosamente como versión {nueva_version}.'
                )
                # Mostrar lista de archivos cargados
                for archivo in archivos_creados:
                    messages.info(self.request, f"📁 {archivo.nombre_original}")
    
        print(f"✅ Proceso completado: {len(archivos_creados)} archivos creados")
        print("=== FIN DEBUG MÚLTIPLES ARCHIVOS ===")
    
        return redirect(self.success_url)
    
    def form_invalid(self, form):
        """Manejo mejorado de formularios inválidos"""
        print("=== FORMULARIO INVÁLIDO ===")
        print(f"Errores del formulario: {form.errors}")
        print(f"Errores no de campo: {form.non_field_errors()}")
        print("===========================")
        
        # Agregar errores como mensajes
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, f"Error: {error}")
                else:
                    messages.error(self.request, f"Error en {field}: {error}")
        
        return super().form_invalid(form)


class ListadoArchivosView(LoginRequiredMixin, ListView):
    """Vista para listar archivos con exportación a Excel"""
    model = Archivo
    template_name = 'archivos/listado_archivos.html'
    context_object_name = 'archivos'
    paginate_by = 20
    
    def get(self, request, *args, **kwargs):
        # VERIFICAR SI ES UNA SOLICITUD DE EXPORTACIÓN
        if request.GET.get('export') == 'excel':
            return self.exportar_excel(request)
        
        return super().get(request, *args, **kwargs)
    
    def get_queryset(self):
        # Obtener perfil del usuario
        try:
            perfil = self.request.user.perfilusuario
            tipo_usuario = perfil.tipo_usuario
        except PerfilUsuario.DoesNotExist:
            messages.warning(self.request, 'Tu usuario no tiene un perfil asignado.')
            return Archivo.objects.none()
        
        # Filtrar por tipo de usuario
        queryset = Archivo.objects.filter(
            fraccion__tipo_usuario_asignado=tipo_usuario
        ).select_related('fraccion', 'usuario').order_by('-created_at')
        
        # FILTRO POR ESTADO (VIGENTE/HISTÓRICO/TODOS)
        estado = self.request.GET.get('estado', 'vigente')  # Por defecto solo vigentes
        
        if estado == 'vigente':
            queryset = queryset.filter(vigente=True)
        elif estado == 'historico':
            queryset = queryset.filter(vigente=False)
        # Si estado == 'todos', no filtramos por vigente
        
        # Filtros adicionales
        fraccion_id = self.request.GET.get('fraccion')
        año = self.request.GET.get('año')
        
        if fraccion_id:
            queryset = queryset.filter(fraccion_id=fraccion_id)
        
        if año:
            queryset = queryset.filter(año=año)
        
        # Funcionalidad de búsqueda
        busqueda = self.request.GET.get('busqueda')
        if busqueda:
            queryset = queryset.filter(
                Q(nombre_original__icontains=busqueda) |
                Q(fraccion__nombre__icontains=busqueda) |
                Q(fraccion__numero__icontains=busqueda)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener fracciones para filtros
        try:
            perfil = self.request.user.perfilusuario
            tipo_usuario = perfil.tipo_usuario
            fracciones = Fraccion.objects.filter(
                tipo_usuario_asignado=tipo_usuario,
                activa=True
            )
        except PerfilUsuario.DoesNotExist:
            fracciones = Fraccion.objects.none()
        
        # Años disponibles
        años = Archivo.objects.filter(
            fraccion__in=fracciones
        ).values_list('año', flat=True).distinct().order_by('-año')
        
        context.update({
            'fracciones': fracciones,
            'años': años,
            'fraccion_seleccionada': self.request.GET.get('fraccion'),
            'año_seleccionado': self.request.GET.get('año'),
            'estado_seleccionado': self.request.GET.get('estado', 'vigente'),
            'busqueda_actual': self.request.GET.get('busqueda', ''),
        })
        
        return context
    
    def exportar_excel(self, request):
        """
        Exporta los archivos filtrados a un archivo Excel agrupado por fracción.
        """
        print("📊 Iniciando exportación a Excel agrupada...")

        queryset = self.get_queryset().order_by(
            'fraccion__numero',
            'año',
            'periodo_especifico',
            '-created_at'
        )

        tipo_usuario_display = self._get_tipo_usuario_display(request)
        wb, ws, styles = self._crear_workbook()
        row_num = self._escribir_encabezados(ws, styles)
        row_num = self._escribir_datos_archivos(ws, queryset, request, row_num, styles)
        self._ajustar_ancho_columnas(ws)
        info_row = self._agregar_informacion_reporte(ws, request, queryset, tipo_usuario_display, row_num)
        self._agregar_estadisticas_fracciones(ws, queryset, info_row)

        response = self._preparar_respuesta_excel(request, queryset, wb)
        print(f"📁 Fracciones incluidas: {len(set(a.fraccion.numero for a in queryset))}")

        return response

    def _get_tipo_usuario_display(self, request):
        try:
            perfil = request.user.perfilusuario
            return perfil.get_tipo_usuario_display()
        except PerfilUsuario.DoesNotExist:
            return "Usuario"

    def _crear_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Archivos Artículo 65"
        styles = {
            'header_font': Font(bold=True, color="FFFFFF", size=12),
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_alignment': Alignment(horizontal="center", vertical="center"),
            'fraccion_font': Font(bold=True, color="FFFFFF", size=11),
            'fraccion_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        return wb, ws, styles

    def _escribir_encabezados(self, ws, styles):
        headers = [
            'Número',
            'Fracción',
            'Año',
            'Tipo Periodo',
            'Archivo',
            'Fecha Carga',
            'Enlace Público'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['header_alignment']
            cell.border = styles['border']
        return 2

    def _escribir_datos_archivos(self, ws, queryset, request, row_num, styles):
        current_fraccion = None
        headers_len = 7
        for archivo in queryset:
            if current_fraccion != archivo.fraccion.numero:
                if current_fraccion is not None:
                    row_num += 1
                ws.merge_cells(f'A{row_num}:G{row_num}')
                title_cell = ws.cell(row=row_num, column=1,
                                    value=f"FRACCIÓN {archivo.fraccion.numero} - {archivo.fraccion.nombre}")
                title_cell.font = styles['fraccion_font']
                title_cell.fill = styles['fraccion_fill']
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in range(1, headers_len + 1):
                    ws.cell(row=row_num, column=col).border = styles['border']
                row_num += 1
                current_fraccion = archivo.fraccion.numero

            enlace_publico = f"{request.scheme}://{request.get_host()}/publico/archivo/{archivo.id}/"
            row_data = [
                archivo.fraccion.numero,
                archivo.fraccion.nombre,
                archivo.año,
                archivo.get_tipo_periodo_display(),
                archivo.nombre_original,
                archivo.created_at.strftime("%d/%m/%Y %H:%M"),
                enlace_publico
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = styles['border']
                if archivo.vigente:
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            row_num += 1
        return row_num

    def _ajustar_ancho_columnas(self, ws):
        column_widths = [
            12,  # Número
            45,  # Fracción
            10,  # Año
            15,  # Tipo Periodo
            40,  # Archivo
            18,  # Fecha Carga
            55   # Enlace Público
        ]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _agregar_informacion_reporte(self, ws, request, queryset, tipo_usuario_display, row_num):
        info_row = row_num + 3
        ws.merge_cells(f'A{info_row}:C{info_row}')
        info_title = ws.cell(row=info_row, column=1, value="INFORMACIÓN DEL REPORTE")
        info_title.font = Font(bold=True, size=11)
        info_title.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        info_row += 1

        info_data = [
            ("Reporte generado por:", f"{request.user.get_full_name() or request.user.username}"),
            ("Fecha de generación:", timezone.now().strftime("%d/%m/%Y %H:%M")),
            ("Tipo de usuario:", tipo_usuario_display),
            ("Total de archivos:", queryset.count()),
        ]
        filtros_info = []
        if request.GET.get('fraccion'):
            try:
                fraccion = Fraccion.objects.get(id=request.GET.get('fraccion'))
                filtros_info.append(f"Fracción: {fraccion.numero} - {fraccion.nombre}")
            except Fraccion.DoesNotExist:
                pass
        if request.GET.get('año'):
            filtros_info.append(f"Año: {request.GET.get('año')}")
        if request.GET.get('estado'):
            estado = request.GET.get('estado')
            filtros_info.append(f"Estado: {estado.capitalize()}")
        if request.GET.get('busqueda'):
            filtros_info.append(f"Búsqueda: {request.GET.get('busqueda')}")
        if filtros_info:
            info_data.append(("Filtros aplicados:", " | ".join(filtros_info)))
        for label, value in info_data:
            ws.cell(row=info_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=info_row, column=2, value=str(value))
            info_row += 1
        return info_row

    def _agregar_estadisticas_fracciones(self, ws, queryset, info_row):
        if queryset.count() > 0:
            stats_fraccion = queryset.values('fraccion__numero', 'fraccion__nombre').annotate(
                total=Count('id')
            ).order_by('fraccion__numero')
            if len(stats_fraccion) > 1:
                info_row += 1
                ws.cell(row=info_row, column=1, value="ARCHIVOS POR FRACCIÓN:").font = Font(bold=True)
                info_row += 1
                for stat in stats_fraccion:
                    ws.cell(row=info_row, column=1, value=f"Fracción {stat['fraccion__numero']}:")
                    ws.cell(row=info_row, column=2, value=f"{stat['total']} archivo(s)")
                    info_row += 1

    def _preparar_respuesta_excel(self, request, queryset, wb):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = timezone.now().strftime("%Y%m%d_%H%M")
        filtro_str = ""
        if request.GET.get('año'):
            filtro_str += f"_{request.GET.get('año')}"
        if request.GET.get('fraccion'):
            try:
                fraccion = Fraccion.objects.get(id=request.GET.get('fraccion'))
                filtro_str += f"_Frac{fraccion.numero}"
            except Fraccion.DoesNotExist:
                pass
        elif queryset.count() > 0:
            filtro_str += "_TodasFracciones"
        filename = f"Archivos_Articulo65{filtro_str}_{fecha_actual}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class HistorialView(LoginRequiredMixin, ListView):
    """Vista para mostrar historial de una fracción"""
    model = Archivo
    template_name = 'archivos/historial.html'
    context_object_name = 'archivos'
    
    def get_queryset(self):
        fraccion_id = self.kwargs['fraccion_id']
        
        # Verificar permisos
        try:
            perfil = self.request.user.perfilusuario
            fraccion = get_object_or_404(Fraccion, id=fraccion_id)
            
            if fraccion.tipo_usuario_asignado != perfil.tipo_usuario:
                messages.error(self.request, 'No tienes permisos para ver el historial de esta fracción.')
                return Archivo.objects.none()
                
        except PerfilUsuario.DoesNotExist:
            messages.error(self.request, 'Tu usuario no tiene un perfil asignado.')
            return Archivo.objects.none()
        
        return Archivo.objects.filter(
            fraccion_id=fraccion_id
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fraccion_id = self.kwargs['fraccion_id']
        context['fraccion'] = get_object_or_404(Fraccion, id=fraccion_id)
        return context


class DescargarArchivoView(LoginRequiredMixin, View):
    """Vista para descargar archivos con URL pública"""
    
    def get(self, request, archivo_id):
        archivo = get_object_or_404(Archivo, id=archivo_id)
        
        # Verificar permisos
        try:
            perfil = request.user.perfilusuario
            if archivo.fraccion.tipo_usuario_asignado != perfil.tipo_usuario:
                raise Http404("No tienes permisos para acceder a este archivo")
        except PerfilUsuario.DoesNotExist:
            raise Http404("Tu usuario no tiene un perfil asignado")
        
        # Registrar acceso
        try:
            HistorialAcceso.objects.create(
                archivo=archivo,
                usuario=request.user,
                ip_address=self.get_client_ip(request)
            )
        except Exception as e:
            print(f"Error al registrar acceso: {e}")
        
        # Servir archivo
        try:
            if archivo.archivo and archivo.archivo.name:
                file_path = safe_join(settings.MEDIA_ROOT, archivo.archivo.name)
                if os.path.exists(file_path):
                    response = FileResponse(
                        open(file_path, 'rb'),
                        as_attachment=True,
                        filename=archivo.nombre_original
                    )
                    return response
                else:
                    raise Http404("Archivo físico no encontrado")
            else:
                raise Http404("Archivo no válido")
        except Exception as e:
            print(f"Error al servir archivo: {e}")
            raise Http404("Error al acceder al archivo")
    
    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip or '0.0.0.0'


class VerArchivoView(LoginRequiredMixin, View):
    """Vista para visualizar archivos en el navegador"""
    
    def get(self, request, archivo_id):
        archivo = get_object_or_404(Archivo, id=archivo_id)
        
        # Verificar permisos (mismo código que DescargarArchivoView)
        try:
            perfil = request.user.perfilusuario
            if archivo.fraccion.tipo_usuario_asignado != perfil.tipo_usuario:
                raise Http404("No tienes permisos para acceder a este archivo")
        except PerfilUsuario.DoesNotExist:
            raise Http404("Tu usuario no tiene un perfil asignado")
        
        # Registrar acceso
        try:
            HistorialAcceso.objects.create(
                archivo=archivo,
                usuario=request.user,
                ip_address=self.get_client_ip(request)
            )
        except Exception as e:
            print(f"Error al registrar acceso: {e}")
        
        # Servir archivo para visualización
        try:
            if archivo.archivo and archivo.archivo.name:
                file_path = safe_join(settings.MEDIA_ROOT, archivo.archivo.name)
                if os.path.exists(file_path):
                    content_type, _ = mimetypes.guess_type(file_path)
                    response = FileResponse(
                        open(file_path, 'rb'),
                        content_type=content_type
                    )
                    return response
                else:
                    raise Http404("Archivo físico no encontrado")
            else:
                raise Http404("Archivo no válido")
        except Exception as e:
            print(f"Error al servir archivo: {e}")
            raise Http404("Error al acceder al archivo")
    
    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip or '0.0.0.0'


class EstadisticasView(LoginRequiredMixin, TemplateView):
    """Vista para mostrar estadísticas"""
    template_name = 'archivos/estadisticas.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener tipo de usuario
        try:
            perfil = self.request.user.perfilusuario
            tipo_usuario = perfil.tipo_usuario
        except PerfilUsuario.DoesNotExist:
            tipo_usuario = None
            messages.warning(self.request, 'Tu usuario no tiene un perfil asignado.')
        
        if tipo_usuario:
            # Estadísticas por fracción
            stats_fraccion = Archivo.objects.filter(
                fraccion__tipo_usuario_asignado=tipo_usuario,
                vigente=True
            ).values(
                'fraccion__numero', 'fraccion__nombre'
            ).annotate(
                total=Count('id')
            ).order_by('fraccion__numero')
            
            # Estadísticas por año
            stats_año = Archivo.objects.filter(
                fraccion__tipo_usuario_asignado=tipo_usuario
            ).values('año').annotate(
                total=Count('id')
            ).order_by('año')
            
            context.update({
                'stats_fraccion': stats_fraccion,
                'stats_año': stats_año,
                'tipo_usuario': tipo_usuario,
            })
        
        return context


class VerArchivoPublicoView(View):
    """Vista PÚBLICA para visualizar archivos sin autenticación"""
    
    def get(self, request, archivo_id):
        archivo = get_object_or_404(Archivo, id=archivo_id)
        
        print(f"📖 Acceso público a archivo: {archivo.nombre_original}")
        print(f"👤 IP: {self.get_client_ip(request)}")
        print(f"🌐 User Agent: {request.META.get('HTTP_USER_AGENT', 'No disponible')}")
        
        # Registrar acceso público (usuario anónimo)
        try:
            HistorialAcceso.objects.create(
                archivo=archivo,
                usuario=None,  # Usuario anónimo
                ip_address=self.get_client_ip(request)
            )
        except Exception as e:
            print(f"⚠️ Error registrando acceso público: {e}")
        
        # Servir archivo para visualización
        try:
            if archivo.archivo and archivo.archivo.name:
                file_path = safe_join(settings.MEDIA_ROOT, archivo.archivo.name)
                if os.path.exists(file_path):
                    content_type, _ = mimetypes.guess_type(file_path)
                    
                    print(f"✅ Sirviendo archivo: {file_path}")
                    print(f"📄 Content-Type: {content_type}")
                    
                    response = FileResponse(
                        open(file_path, 'rb'),
                        content_type=content_type,
                        filename=archivo.nombre_original
                    )
                    
                    # Headers adicionales para mejor experiencia
                    response['Content-Disposition'] = f'inline; filename="{archivo.nombre_original}"'
                    response['X-Frame-Options'] = 'SAMEORIGIN'
                    
                    return response
                else:
                    print(f"❌ Archivo físico no encontrado: {file_path}")
                    raise Http404("Archivo no encontrado")
            else:
                print("❌ Archivo no válido")
                raise Http404("Archivo no válido")
                
        except Exception as e:
            print(f"❌ Error sirviendo archivo público: {e}")
            raise Http404("Error al acceder al archivo")
    
    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip or '0.0.0.0'


class DescargarArchivoPublicoView(View):
    """Vista PÚBLICA para descargar archivos sin autenticación"""
    
    def get(self, request, archivo_id):
        archivo = get_object_or_404(Archivo, id=archivo_id)
        
        print(f"⬇️ Descarga pública de archivo: {archivo.nombre_original}")
        print(f"👤 IP: {self.get_client_ip(request)}")
        
        # Registrar acceso público
        try:
            HistorialAcceso.objects.create(
                archivo=archivo,
                usuario=None,  # Usuario anónimo
                ip_address=self.get_client_ip(request)
            )
        except Exception as e:
            print(f"⚠️ Error registrando descarga pública: {e}")
        
        # Servir archivo para descarga
        try:
            if archivo.archivo and archivo.archivo.name:
                file_path = safe_join(settings.MEDIA_ROOT, archivo.archivo.name)
                if os.path.exists(file_path):
                    print(f"✅ Descargando archivo: {file_path}")
                    
                    response = FileResponse(
                        open(file_path, 'rb'),
                        as_attachment=True,
                        filename=archivo.nombre_original
                    )
                    
                    # Headers adicionales
                    response['Content-Length'] = archivo.tamaño
                    
                    return response
                else:
                    print(f"❌ Archivo físico no encontrado para descarga: {file_path}")
                    raise Http404("Archivo no encontrado")
            else:
                print("❌ Archivo no válido")
                raise Http404("Archivo no válido")
                
        except Exception as e:
            print(f"❌ Error en descarga pública: {e}")
            raise Http404("Error al descargar el archivo")
    
    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip or '0.0.0.0'